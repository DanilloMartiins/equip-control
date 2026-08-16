from io import BytesIO
from datetime import datetime, date

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify, session
from werkzeug.security import check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook

from config import REGIONAIS, LOGIN_USER, LOGIN_PASS_HASH
from db import get_db, padronizar_tipo, IntegrityError, add_envio, add_pendencia, add_historico, USING_PG
from matcher import identificar_colunas, mapear_para_insert

bp = Blueprint('routes', __name__)

@bp.before_request
def exigir_login():
    if session.get('usuario') != LOGIN_USER:
        if request.endpoint != 'routes.login':
            return redirect(url_for('routes.login'))

# -----------------------------------------------------------------------
# LOGIN / LOGOUT
# -----------------------------------------------------------------------
@bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form.get('usuario', '').strip()
        senha = request.form.get('senha', '')
        if usuario == LOGIN_USER and check_password_hash(LOGIN_PASS_HASH, senha):
            session['usuario'] = usuario
            flash(f'Bem-vindo, {usuario}!', 'success')
            return redirect(url_for('routes.index'))
        flash('Usuário ou senha inválidos.', 'danger')
    return render_template('login.html')

@bp.route('/logout')
def logout():
    session.pop('usuario', None)
    flash('Você saiu do sistema.', 'info')
    return redirect(url_for('routes.login'))

# -----------------------------------------------------------------------
# PERFIL
# -----------------------------------------------------------------------
@bp.route('/perfil', methods=['GET', 'POST'])
def perfil():
    conn = get_db()

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip() or None
        foto = request.form.get('foto') or None

        if nome or foto:
            if USING_PG:
                conn.execute("""
                    INSERT INTO usuarios (usuario, nome, foto, updated_at)
                    VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
                    ON CONFLICT (usuario) DO UPDATE SET
                        nome = EXCLUDED.nome,
                        foto = COALESCE(EXCLUDED.foto, usuarios.foto),
                        updated_at = CURRENT_TIMESTAMP
                """, (session.get('usuario'), nome, foto))
            else:
                conn.execute("""
                    INSERT INTO usuarios (usuario, nome, foto)
                    VALUES (?, ?, ?)
                    ON CONFLICT (usuario) DO UPDATE SET
                        nome = excluded.nome,
                        foto = COALESCE(excluded.foto, usuarios.foto)
                """, (session.get('usuario'), nome, foto))
            conn.commit()
        conn.close()
        flash('Perfil alterado com sucesso!', 'success')
        return redirect(url_for('routes.perfil'))

    perfil_row = None
    if session.get('usuario'):
        perfil_row = conn.execute(
            "SELECT nome, foto FROM usuarios WHERE usuario = %s", (session['usuario'],)
        ).fetchone()
    conn.close()

    return render_template('perfil.html',
        perfil_nome=perfil_row['nome'] if perfil_row else None,
        perfil_foto=perfil_row['foto'] if perfil_row else None)

# -----------------------------------------------------------------------
# DASHBOARD
# -----------------------------------------------------------------------
@bp.route('/')
def index() -> str:
    conn = get_db()
    total = conn.execute("SELECT COUNT(*) AS total FROM equipamentos").fetchone()['total']
    por_regional = conn.execute("""
        SELECT regional, COUNT(*) as qtd FROM equipamentos
        GROUP BY regional ORDER BY qtd DESC
    """).fetchall()
    por_tipo = conn.execute("""
        SELECT tipo, COUNT(*) as qtd FROM equipamentos
        GROUP BY tipo ORDER BY qtd DESC
    """).fetchall()
    pendentes = conn.execute("""
        SELECT COUNT(*) AS total FROM equipamentos WHERE status = 'pendente'
    """).fetchone()['total']
    concluidos = conn.execute("""
        SELECT COUNT(*) AS total FROM equipamentos WHERE status = 'concluido'
    """).fetchone()['total']
    recentes = conn.execute("""
        SELECT * FROM equipamentos ORDER BY created_at DESC LIMIT 10
    """).fetchall()
    conn.close()

    return render_template('index.html',
        total=total, por_regional=por_regional, por_tipo=por_tipo,
        pendentes=pendentes, concluidos=concluidos, recentes=recentes, regionais=REGIONAIS)

# -----------------------------------------------------------------------
# CADASTRAR
# -----------------------------------------------------------------------
@bp.route('/cadastrar', methods=['GET', 'POST'])
def cadastrar() -> str:
    if request.method == 'POST':
        codigo = request.form['codigo'].strip()
        regional = request.form['regional']
        tipo = padronizar_tipo(request.form['tipo'])
        fabricante = request.form.get('fabricante') or None
        modelo = request.form.get('modelo') or None
        numero_serie = request.form.get('numero_serie') or None
        local_instalacao = request.form.get('local_instalacao') or None
        idbdit = request.form.get('idbdit') or None
        origem = request.form.get('origem') or 'oficio'
        data_cad = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        data_entrada = request.form.get('data_entrada_operacao') or None
        data_sol = request.form.get('data_solicitacao') or None

        if not codigo or not regional or not tipo:
            flash('Preencha todos os campos obrigatórios.', 'danger')
            return redirect(url_for('routes.cadastrar'))

        conn = get_db()
        try:
            # Mensagens de confirmação seguem o padrão "Entidade + verbo + com sucesso"
            # (ex: "Equipamento cadastrado com sucesso!") de propósito: deixam a porta
            # aberta pro sistema escalar e gerenciar outras entidades (pessoa, ativo,
            # veículo etc.) no futuro, sem precisar reescrever nada.
            conn.execute(
                "INSERT INTO equipamentos (codigo, regional, tipo, fabricante, modelo, numero_serie, local_instalacao, idbdit, origem, data_cadastro, data_entrada_operacao, data_solicitacao) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (codigo, regional, tipo, fabricante, modelo, numero_serie, local_instalacao, idbdit, origem, data_cad, data_entrada, data_sol)
            )
            conn.commit()
            flash('Equipamento cadastrado com sucesso!', 'success')
            return redirect(url_for('routes.index'))
        except IntegrityError:
            flash(f'Código {codigo} já existe no sistema.', 'danger')
        finally:
            conn.close()

    return render_template('cadastrar.html', regionais=REGIONAIS, hoje=date.today().isoformat())

# -----------------------------------------------------------------------
# IMPORTAR EXCEL
# -----------------------------------------------------------------------
@bp.route('/importar', methods=['GET', 'POST'])
def importar() -> str:
    if request.method == 'POST':
        if 'arquivo' not in request.files:
            flash('Nenhum arquivo enviado.', 'danger')
            return redirect(url_for('routes.importar'))

        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            flash('Selecione um arquivo.', 'danger')
            return redirect(url_for('routes.importar'))

        try:
            wb = load_workbook(arquivo)
            ws = wb.active

            importados = 0
            ignorados = 0
            conn = get_db()

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                raw_id = row[0]
                raw_regional = str(row[1]).strip() if row[1] else ''
                raw_tipo = str(row[2]).strip() if row[2] else ''
                raw_cad = row[3]
                raw_env = row[4]

                if raw_id is None or raw_regional.lower() not in [r.lower() for r in REGIONAIS]:
                    ignorados += 1
                    continue

                codigo = str(int(raw_id)) if isinstance(raw_id, float) else str(raw_id)
                regional = raw_regional
                tipo = padronizar_tipo(raw_tipo)

                data_entrada = str(raw_cad)[:10] if raw_cad else None
                data_env = str(raw_env)[:10] if raw_env else None

                try:
                    conn.execute(
                        "INSERT INTO equipamentos (codigo, regional, tipo, data_entrada_operacao, data_solicitacao) VALUES (%s, %s, %s, %s, %s)",
                        (codigo, regional, tipo, data_entrada, data_env)
                    )
                    importados += 1
                except IntegrityError:
                    ignorados += 1

            conn.commit()
            conn.close()
            flash(f'Importação concluída! {importados} importados, {ignorados} ignorados.', 'success')
            return redirect(url_for('routes.index'))
        except Exception as e:
            flash(f'Erro ao importar: {str(e)}', 'danger')
            return redirect(url_for('routes.importar'))

    return render_template('importar.html')

# -----------------------------------------------------------------------
# RELATORIO WEB
# -----------------------------------------------------------------------
LIMITES_RELATORIO = ['50', '100', '500', '1000', 'todos']

def _filtros_query():
    busca = request.args.get('q', '').strip()
    regionais = [r for r in request.args.getlist('regional') if r]
    tipos = [t for t in request.args.getlist('tipo') if t]
    status = [s for s in request.args.getlist('status') if s]
    return busca, regionais, tipos, status

def _where_filtros(busca, regionais, tipos, status):
    wheres = []
    params = []
    if busca:
        wheres.append("(LOWER(codigo) LIKE LOWER(%s) OR LOWER(regional) LIKE LOWER(%s) OR LOWER(tipo) LIKE LOWER(%s))")
        params.extend([f'%{busca}%'] * 3)
    if regionais:
        wheres.append("regional IN (" + ','.join(['%s'] * len(regionais)) + ")")
        params.extend(regionais)
    if tipos:
        wheres.append("tipo IN (" + ','.join(['%s'] * len(tipos)) + ")")
        params.extend(tipos)
    if status:
        wheres.append("status IN (" + ','.join(['%s'] * len(status)) + ")")
        params.extend(status)
    return wheres, params

def _qs_filtros(busca, regionais, tipos, status):
    from urllib.parse import quote
    parts = []
    if busca:
        parts.append('q=' + quote(busca))
    for r in regionais:
        parts.append('regional=' + quote(r))
    for t in tipos:
        parts.append('tipo=' + quote(t))
    for s in status:
        parts.append('status=' + quote(s))
    return '&'.join(parts)

@bp.route('/relatorio')
def relatorio() -> str:
    limite = request.args.get('limite', '50')
    if limite not in LIMITES_RELATORIO:
        limite = '50'

    busca, regionais, tipos, status = _filtros_query()
    wheres, params = _where_filtros(busca, regionais, tipos, status)
    where_clause = " WHERE " + " AND ".join(wheres) if wheres else ""

    conn = get_db()

    total = conn.execute("SELECT COUNT(*) AS total FROM equipamentos").fetchone()['total']
    total_filtro = conn.execute(
        f"SELECT COUNT(*) AS total FROM equipamentos{where_clause}", params
    ).fetchone()['total']

    tipos_opcoes = [t['tipo'] for t in conn.execute(
        "SELECT DISTINCT tipo FROM equipamentos WHERE tipo IS NOT NULL AND tipo != '' ORDER BY tipo"
    ).fetchall()]
    status_opcoes = [s['status'] for s in conn.execute(
        "SELECT DISTINCT status FROM equipamentos WHERE status IS NOT NULL AND status != '' ORDER BY status"
    ).fetchall()]

    if limite == 'todos':
        equipamentos = conn.execute(
            f"SELECT * FROM equipamentos{where_clause} ORDER BY regional, tipo, codigo", params
        ).fetchall()
        pagina_atual = 1
        total_paginas = 1
    else:
        limite_int = int(limite)
        total_paginas = max((total_filtro + limite_int - 1) // limite_int, 1)
        pagina_atual = max(request.args.get('pagina', 1, type=int), 1)
        if pagina_atual > total_paginas:
            pagina_atual = total_paginas
        offset = (pagina_atual - 1) * limite_int
        params_pag = params + [limite_int, offset]
        equipamentos = conn.execute(
            f"SELECT * FROM equipamentos{where_clause} ORDER BY regional, tipo, codigo LIMIT %s OFFSET %s",
            params_pag
        ).fetchall()

    por_regional = conn.execute("""
        SELECT regional, COUNT(*) as qtd FROM equipamentos GROUP BY regional ORDER BY qtd DESC
    """).fetchall()
    por_tipo = conn.execute("""
        SELECT tipo, COUNT(*) as qtd FROM equipamentos GROUP BY tipo ORDER BY qtd DESC
    """).fetchall()
    conn.close()

    filtros_qs = _qs_filtros(busca, regionais, tipos, status)

    return render_template('relatorio.html',
        equipamentos=equipamentos, por_regional=por_regional, por_tipo=por_tipo,
        total=total, total_filtro=total_filtro, limite=limite, LIMITES_RELATORIO=LIMITES_RELATORIO,
        pagina_atual=pagina_atual, total_paginas=total_paginas,
        busca=busca, regionais=regionais, tipos=tipos, status=status,
        tipos_opcoes=tipos_opcoes, status_opcoes=status_opcoes, filtros_qs=filtros_qs)

# -----------------------------------------------------------------------
# EXPORTAR EXCEL
# -----------------------------------------------------------------------
@bp.route('/exportar')
def exportar():
    limite = request.args.get('limite', 'todos')
    if limite not in LIMITES_RELATORIO:
        limite = 'todos'

    busca, regionais, tipos, status = _filtros_query()
    wheres, params = _where_filtros(busca, regionais, tipos, status)
    where_clause = " WHERE " + " AND ".join(wheres) if wheres else ""

    conn = get_db()

    if limite == 'todos':
        dados = conn.execute("""
            SELECT codigo, regional, tipo, data_entrada_operacao, data_solicitacao
            FROM equipamentos{where_clause} ORDER BY regional, tipo, codigo
        """.format(where_clause=where_clause), params).fetchall()
    else:
        limite_int = int(limite)
        pagina_atual = max(request.args.get('pagina', 1, type=int), 1)
        offset = (pagina_atual - 1) * limite_int
        params_pag = params + [limite_int, offset]
        dados = conn.execute("""
            SELECT codigo, regional, tipo, data_entrada_operacao, data_solicitacao
            FROM equipamentos{where_clause} ORDER BY regional, tipo, codigo LIMIT %s OFFSET %s
        """.format(where_clause=where_clause), params_pag).fetchall()

    por_regional = conn.execute("""
        SELECT regional, COUNT(*) as qtd FROM equipamentos GROUP BY regional ORDER BY qtd DESC
    """).fetchall()
    por_tipo = conn.execute("""
        SELECT tipo, COUNT(*) as qtd FROM equipamentos GROUP BY tipo ORDER BY qtd DESC
    """).fetchall()
    total = conn.execute("SELECT COUNT(*) AS total FROM equipamentos").fetchone()['total']
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipamentos Cadastrados"

    AZUL_E = "1F4E79"
    AZUL_M = "2E75B6"
    thin = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    fill_titulo = PatternFill(start_color=AZUL_E, end_color=AZUL_E, fill_type="solid")
    fill_header = PatternFill(start_color=AZUL_M, end_color=AZUL_M, fill_type="solid")
    fill_l1 = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    fill_l2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    ca = Alignment(horizontal="center", vertical="center")
    cl = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A1:E1")
    ws["A1"].value = "RELATÓRIO DE EQUIPAMENTOS CADASTRADOS - Equip-Control"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws["A1"].alignment = ca
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:E2")
    ws["A2"].value = f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=10, color="B4C6E7")
    ws["A2"].fill = fill_titulo
    ws["A2"].alignment = ca
    ws.row_dimensions[2].height = 22

    card_fill = PatternFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")
    card_border = Border(
        left=Side(style="thin", color="D0D8E8"),
        right=Side(style="thin", color="D0D8E8"),
        top=Side(style="thin", color="D0D8E8"),
        bottom=Side(style="thin", color="D0D8E8"),
    )
    ws.merge_cells("A4:B4")
    ws["A4"].value = "Total de Equipamentos"
    ws["A4"].font = Font(name="Calibri", size=9, color="666666")
    ws["A4"].alignment = ca
    ws.merge_cells("A5:B5")
    ws["A5"].value = total
    ws["A5"].font = Font(name="Calibri", size=14, bold=True, color=AZUL_E)
    ws["A5"].alignment = ca

    ws.merge_cells("C4:C4")
    ws["C4"].value = "Regionais"
    ws["C4"].font = Font(name="Calibri", size=9, color="666666")
    ws["C4"].alignment = ca
    ws["C5"].value = len(por_regional)
    ws["C5"].font = Font(name="Calibri", size=14, bold=True, color=AZUL_E)
    ws["C5"].alignment = ca

    ws.merge_cells("D4:E4")
    ws["D4"].value = "Tipos de Equipamento"
    ws["D4"].font = Font(name="Calibri", size=9, color="666666")
    ws["D4"].alignment = ca
    ws.merge_cells("D5:E5")
    ws["D5"].value = len(por_tipo)
    ws["D5"].font = Font(name="Calibri", size=14, bold=True, color=AZUL_E)
    ws["D5"].alignment = ca

    for col_idx in range(1, 6):
        for row_idx in [4, 5]:
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = card_fill
            c.border = card_border
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 36

    h_row = 7
    headers = ["Nº Equipamento", "Regional Axia", "Tipo de Equipamento",
               "Data de Cadastro", "Data de Solicitação de Envio"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=h_row, column=col, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = fill_header
        c.alignment = ca
        c.border = thin
    ws.row_dimensions[h_row].height = 30

    for i, d in enumerate(dados):
        r = h_row + 1 + i
        fill = fill_l1 if i % 2 == 0 else fill_l2
        vals = [d['codigo'], d['regional'], d['tipo'], d['data_entrada_operacao'], d['data_solicitacao']]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = Font(name="Calibri", size=10, color="333333")
            c.fill = fill
            c.border = thin
            c.alignment = ca if col_idx != 3 else cl
        ws.row_dimensions[r].height = 22

    final_row = h_row + 1 + len(dados) + 1
    ws.merge_cells(f"A{final_row}:E{final_row}")
    ws.cell(row=final_row, column=1, value="RESUMO POR REGIONAL")
    ws.cell(row=final_row, column=1).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws.cell(row=final_row, column=1).fill = fill_titulo
    ws.cell(row=final_row, column=1).alignment = ca
    for col_idx in range(2, 6):
        ws.cell(row=final_row, column=col_idx).fill = fill_titulo
        ws.cell(row=final_row, column=col_idx).border = thin
    ws.row_dimensions[final_row].height = 24

    r = final_row + 1
    bg_reg = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
    for row_reg in por_regional:
        ws.cell(row=r, column=1, value="").border = thin
        ws.cell(row=r, column=1).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        ws.cell(row=r, column=2, value=row_reg['regional'])
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, bold=True, color=AZUL_E)
        ws.cell(row=r, column=2).alignment = ca
        ws.cell(row=r, column=2).border = thin
        ws.cell(row=r, column=2).fill = bg_reg
        ws.merge_cells(f"C{r}:D{r}")
        ws.cell(row=r, column=3, value=f"{row_reg['qtd']} equipamento(s)")
        ws.cell(row=r, column=3).font = Font(name="Calibri", size=10, color="333333")
        ws.cell(row=r, column=3).alignment = cl
        ws.cell(row=r, column=3).border = thin
        ws.cell(row=r, column=3).fill = bg_reg
        ws.cell(row=r, column=5, value=row_reg['qtd'])
        ws.cell(row=r, column=5).font = Font(name="Calibri", size=10, bold=True, color=AZUL_E)
        ws.cell(row=r, column=5).alignment = ca
        ws.cell(row=r, column=5).border = thin
        ws.cell(row=r, column=5).fill = bg_reg
        ws.row_dimensions[r].height = 22
        r += 1

    for col_idx in range(1, 6):
        ws.cell(row=r, column=col_idx).border = thin
        ws.cell(row=r, column=col_idx).fill = fill_header
    ws.cell(row=r, column=1, value="")
    ws.cell(row=r, column=2, value="TOTAL GERAL")
    ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws.cell(row=r, column=2).alignment = ca
    ws.merge_cells(f"C{r}:D{r}")
    ws.cell(row=r, column=3, value=f"{total} equipamento(s) cadastrado(s)")
    ws.cell(row=r, column=3).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws.cell(row=r, column=3).alignment = cl
    ws.cell(row=r, column=5, value=total)
    ws.cell(row=r, column=5).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.cell(row=r, column=5).alignment = ca
    ws.row_dimensions[r].height = 24
    r += 2

    ws.merge_cells(f"A{r}:E{r}")
    ws.cell(row=r, column=1, value="RESUMO POR TIPO DE EQUIPAMENTO")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws.cell(row=r, column=1).fill = fill_titulo
    ws.cell(row=r, column=1).alignment = ca
    for col_idx in range(2, 6):
        ws.cell(row=r, column=col_idx).fill = fill_titulo
        ws.cell(row=r, column=col_idx).border = thin
    ws.row_dimensions[r].height = 24
    r += 1

    bg_tipo = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    for row_tp in por_tipo:
        ws.cell(row=r, column=1, value="").border = thin
        ws.cell(row=r, column=1).fill = bg_tipo
        ws.merge_cells(f"B{r}:C{r}")
        ws.cell(row=r, column=2, value=row_tp['tipo'])
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, color="333333")
        ws.cell(row=r, column=2).alignment = cl
        ws.cell(row=r, column=2).border = thin
        ws.cell(row=r, column=2).fill = bg_tipo
        ws.cell(row=r, column=3).border = thin
        ws.cell(row=r, column=5, value=row_tp['qtd'])
        ws.cell(row=r, column=5).font = Font(name="Calibri", size=10, bold=True, color=AZUL_E)
        ws.cell(row=r, column=5).alignment = ca
        ws.cell(row=r, column=5).border = thin
        ws.cell(row=r, column=5).fill = bg_tipo
        ws.row_dimensions[r].height = 22
        r += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 24
    ws.sheet_properties.tabColor = AZUL_M
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
        download_name=f"Relatorio_Equipamentos_EquipControl_{datetime.now().strftime('%d%m%Y')}.xlsx",
        as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# -----------------------------------------------------------------------
# LISTAGEM COM FILTRO
# -----------------------------------------------------------------------
@bp.route('/equipamentos')
def listar_equipamentos() -> str:
    status_filtro = request.args.get('status', '')
    busca = request.args.get('q', '').strip()
    conn = get_db()

    wheres = []
    params = []

    if status_filtro:
        wheres.append("status = %s")
        params.append(status_filtro)

    if busca:
        wheres.append("codigo ILIKE %s")
        params.append(f'%{busca}%')

    where_clause = " WHERE " + " AND ".join(wheres) if wheres else ""

    equipamentos = conn.execute(
        f"SELECT * FROM equipamentos{where_clause} ORDER BY regional, tipo, codigo",
        params
    ).fetchall()
    total_filtro = conn.execute(
        f"SELECT COUNT(*) AS total FROM equipamentos{where_clause}",
        params
    ).fetchone()['total']

    por_regional = conn.execute("""
        SELECT regional, COUNT(*) as qtd FROM equipamentos GROUP BY regional ORDER BY qtd DESC
    """).fetchall()
    por_tipo = conn.execute("""
        SELECT tipo, COUNT(*) as qtd FROM equipamentos GROUP BY tipo ORDER BY qtd DESC
    """).fetchall()
    total = conn.execute("SELECT COUNT(*) AS total FROM equipamentos").fetchone()['total']
    conn.close()

    return render_template('equipamentos.html',
        equipamentos=equipamentos, por_regional=por_regional, por_tipo=por_tipo,
        total=total, total_filtro=total_filtro, status_filtro=status_filtro, busca=busca)

# -----------------------------------------------------------------------
# API E DELETAR
# -----------------------------------------------------------------------
@bp.route('/api/equipamentos')
def api_equipamentos():
    conn = get_db()
    dados = conn.execute("SELECT * FROM equipamentos ORDER BY created_at DESC").fetchall()
    conn.close()
    return jsonify([dict(d) for d in dados])

# -----------------------------------------------------------------------
# IMPORT INTELIGENTE (FUZZY MATCH)
# -----------------------------------------------------------------------
@bp.route('/importar-smart', methods=['GET', 'POST'])
def importar_smart() -> str:
    if request.method == 'POST':
        if 'arquivo' not in request.files:
            flash('Nenhum arquivo enviado.', 'danger')
            return redirect(url_for('routes.importar_smart'))

        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            flash('Selecione um arquivo.', 'danger')
            return redirect(url_for('routes.importar_smart'))

        try:
            wb = load_workbook(arquivo, read_only=True, data_only=True)
            ws = wb.active

            cabecalhos = []
            for row in ws.iter_rows(max_row=20, values_only=True):
                vals = [str(v).strip() if v else '' for v in row]
                vals = [v for v in vals if v]
                if len(vals) >= 3:
                    cabecalhos = vals
                    break

            mapping = identificar_colunas(cabecalhos)

            if 'codigo' not in mapping:
                flash('Não foi possível identificar a coluna de código do equipamento na planilha.', 'danger')
                return redirect(url_for('routes.importar_smart'))

            conn = get_db()
            importados = 0
            ignorados = 0
            erros = []

            colunas_db = ['regional', 'tipo', 'fabricante', 'modelo', 'numero_serie',
                          'local_instalacao', 'idbdit', 'data_entrada_operacao', 'data_solicitacao']

            for row in ws.iter_rows(min_row=2, values_only=True):
                dados = mapear_para_insert(row, mapping, cabecalhos)

                codigo = dados.get('codigo')
                regional = dados.get('regional') or ''
                tipo = dados.get('tipo') or ''

                if not codigo or not tipo:
                    ignorados += 1
                    continue

                regional = regional.split('|', 1)[-1].strip() if '|' in regional else regional.strip()
                tipo = padronizar_tipo(tipo)
                data_entrada = dados.get('data_entrada_operacao') or dados.get('data_cadastro')
                data_sol = dados.get('data_solicitacao')

                try:
                    cur = conn.execute(
                        "SELECT id, status FROM equipamentos WHERE codigo = %s", (codigo,)
                    )
                    existente = cur.fetchone()

                    if existente:
                        sets = []
                        params = []
                        for col in colunas_db:
                            val = dados.get(col)
                            if val:
                                sets.append(f"{col} = %s")
                                params.append(val)
                        if existente['status'] == 'pendente':
                            sets.append("status = 'concluido'")
                        if sets:
                            params.append(codigo)
                            conn.execute(
                                f"UPDATE equipamentos SET {', '.join(sets)} WHERE codigo = %s",
                                params
                            )
                        importados += 1
                    else:
                        conn.execute("""
                            INSERT INTO equipamentos
                                (codigo, regional, tipo, fabricante, modelo, numero_serie,
                                 local_instalacao, idbdit, data_entrada_operacao, data_solicitacao)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            codigo, regional, tipo,
                            dados.get('fabricante'), dados.get('modelo'),
                            dados.get('numero_serie'), dados.get('local_instalacao'),
                            dados.get('idbdit'), data_entrada, data_sol
                        ))
                        importados += 1
                except IntegrityError:
                    ignorados += 1
                except Exception as e:
                    erros.append(str(e)[:80])

            conn.commit()
            conn.close()
            wb.close()

            msg = f'Importação concluída! {importados} importados, {ignorados} ignorados.'
            if erros:
                msg += f' {len(erros)} erro(s).'
            flash(msg, 'success')
            return redirect(url_for('routes.index'))

        except Exception as e:
            flash(f'Erro ao processar planilha: {str(e)}', 'danger')
            return redirect(url_for('routes.importar_smart'))

    return render_template('importar_smart.html')

@bp.route('/deletar/<int:id>', methods=['POST'])
def deletar(id: int) -> str:
    conn = get_db()
    conn.execute("DELETE FROM equipamentos WHERE id = %s", (id,))
    conn.commit()
    conn.close()
    flash('Equipamento deletado com sucesso!', 'success')
    return redirect(url_for('routes.index'))

# -----------------------------------------------------------------------
# DETALHE DO EQUIPAMENTO
# -----------------------------------------------------------------------
@bp.route('/equipamento/<int:id>')
def equipamento_detalhe(id: int) -> str:
    conn = get_db()
    eq = conn.execute("SELECT * FROM equipamentos WHERE id = %s", (id,)).fetchone()
    if not eq:
        conn.close()
        flash('Equipamento não encontrado.', 'danger')
        return redirect(url_for('routes.index'))

    historico = conn.execute(
        "SELECT * FROM historico WHERE equipamento_id = %s ORDER BY data_ocorrencia DESC, id DESC",
        (id,)
    ).fetchall()
    envios = conn.execute(
        "SELECT * FROM envios WHERE equipamento_id = %s ORDER BY data_envio DESC, id DESC",
        (id,)
    ).fetchall()
    pendencias = conn.execute(
        "SELECT * FROM pendencias WHERE equipamento_id = %s ORDER BY data_pendencia DESC, id DESC",
        (id,)
    ).fetchall()
    conn.close()

    return render_template('equipamento.html',
        eq=eq, historico=historico, envios=envios, pendencias=pendencias)

# -----------------------------------------------------------------------
# EDITAR EQUIPAMENTO
# -----------------------------------------------------------------------
@bp.route('/equipamento/<int:id>/editar', methods=['GET', 'POST'])
def equipamento_editar(id: int) -> str:
    conn = get_db()
    if request.method == 'POST':
        dados = {k: request.form.get(k) or None for k in [
            'codigo', 'regional', 'tipo', 'fabricante', 'modelo',
            'numero_serie', 'local_instalacao', 'idbdit', 'origem',
            'status', 'data_entrada_operacao', 'data_solicitacao'
        ]}
        dados['tipo'] = padronizar_tipo(dados['tipo'])
        dados['data_cadastro'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute("""
            UPDATE equipamentos SET
                codigo=%s, regional=%s, tipo=%s, fabricante=%s, modelo=%s,
                numero_serie=%s, local_instalacao=%s, idbdit=%s, origem=%s,
                status=%s, data_entrada_operacao=%s, data_cadastro=%s, data_solicitacao=%s
            WHERE id=%s
        """, (*dados.values(), id))
        add_historico(conn, id, 'edicao', 'Equipamento editado')
        conn.commit()
        conn.close()
        flash('Equipamento alterado com sucesso!', 'success')
        return redirect(url_for('routes.equipamento_detalhe', id=id))

    eq = conn.execute("SELECT * FROM equipamentos WHERE id = %s", (id,)).fetchone()
    conn.close()
    if not eq:
        flash('Equipamento não encontrado.', 'danger')
        return redirect(url_for('routes.index'))
    return render_template('editar.html', eq=eq, regionais=REGIONAIS)

# -----------------------------------------------------------------------
# ADICIONAR ENVIO
# -----------------------------------------------------------------------
@bp.route('/equipamento/<int:id>/envio', methods=['POST'])
def equipamento_envio(id: int) -> str:
    id_envio = request.form['id_envio'].strip()
    destino = request.form['destino'].strip()
    data_envio = request.form.get('data_envio') or None
    observacao = request.form.get('observacao') or None

    if not id_envio or not destino:
        flash('Preencha ID do envio e destino.', 'danger')
        return redirect(url_for('routes.equipamento_detalhe', id=id))

    conn = get_db()
    add_envio(conn, id, id_envio, destino, data_envio, observacao)
    conn.execute("UPDATE equipamentos SET status = 'enviado' WHERE id = %s", (id,))
    conn.commit()
    conn.close()
    flash('Envio registrado!', 'success')
    return redirect(url_for('routes.equipamento_detalhe', id=id))

# -----------------------------------------------------------------------
# ADICIONAR PENDENCIA
# -----------------------------------------------------------------------
@bp.route('/equipamento/<int:id>/pendencia', methods=['POST'])
def equipamento_pendencia(id: int) -> str:
    motivo = request.form['motivo'].strip()
    origem = request.form.get('origem') or None
    data_pendencia = request.form.get('data_pendencia') or None

    if not motivo:
        flash('Descreva o motivo da pendência.', 'danger')
        return redirect(url_for('routes.equipamento_detalhe', id=id))

    conn = get_db()
    add_pendencia(conn, id, motivo, origem, data_pendencia)
    conn.execute("UPDATE equipamentos SET status = 'pendente' WHERE id = %s", (id,))
    conn.commit()
    conn.close()
    flash('Pendência registrada!', 'success')
    return redirect(url_for('routes.equipamento_detalhe', id=id))

# -----------------------------------------------------------------------
# RESOLVER PENDENCIA
# -----------------------------------------------------------------------
@bp.route('/pendencia/<int:id>/resolver', methods=['POST'])
def pendencia_resolver(id: int) -> str:
    conn = get_db()
    cur = conn.execute("SELECT equipamento_id FROM pendencias WHERE id = %s", (id,))
    pend = cur.fetchone()
    if pend:
        from db import resolve_pendencia
        resolve_pendencia(conn, id)
        conn.commit()
        flash('Pendência resolvida!', 'success')
        conn.close()
        return redirect(url_for('routes.equipamento_detalhe', id=pend['equipamento_id']))
    conn.close()
    flash('Pendência não encontrada.', 'danger')
    return redirect(url_for('routes.index'))

# -----------------------------------------------------------------------
# ADICIONAR OBSERVACAO NO HISTORICO
# -----------------------------------------------------------------------
@bp.route('/equipamento/<int:id>/observacao', methods=['POST'])
def equipamento_observacao(id: int) -> str:
    descricao = request.form['descricao'].strip()
    data_ocorrencia = request.form.get('data_ocorrencia') or None
    if not descricao:
        flash('Escreva uma observação.', 'danger')
        return redirect(url_for('routes.equipamento_detalhe', id=id))
    conn = get_db()
    add_historico(conn, id, 'observacao', descricao, data_ocorrencia)
    conn.commit()
    conn.close()
    flash('Observação adicionada!', 'success')
    return redirect(url_for('routes.equipamento_detalhe', id=id))

# -----------------------------------------------------------------------
# LISTAR ENVIOS
# -----------------------------------------------------------------------
@bp.route('/envios')
def listar_envios() -> str:
    conn = get_db()
    envios = conn.execute("""
        SELECT e.*, eq.codigo as eq_codigo, eq.regional as eq_regional, eq.tipo as eq_tipo
        FROM envios e
        JOIN equipamentos eq ON eq.id = e.equipamento_id
        ORDER BY e.created_at DESC LIMIT 100
    """).fetchall()
    conn.close()
    return render_template('envios.html', envios=envios)

# -----------------------------------------------------------------------
# LISTAR PENDENCIAS
# -----------------------------------------------------------------------
@bp.route('/pendencias')
def listar_pendencias() -> str:
    conn = get_db()
    ativas = conn.execute("""
        SELECT p.*, eq.codigo as eq_codigo, eq.regional as eq_regional, eq.tipo as eq_tipo
        FROM pendencias p
        JOIN equipamentos eq ON eq.id = p.equipamento_id
        WHERE p.resolvida = 0
        ORDER BY p.data_pendencia DESC
    """).fetchall()
    resolvidas = conn.execute("""
        SELECT p.*, eq.codigo as eq_codigo, eq.regional as eq_regional, eq.tipo as eq_tipo
        FROM pendencias p
        JOIN equipamentos eq ON eq.id = p.equipamento_id
        WHERE p.resolvida = 1
        ORDER BY p.data_resolucao DESC LIMIT 50
    """).fetchall()
    conn.close()
    return render_template('pendencias.html', ativas=ativas, resolvidas=resolvidas)
